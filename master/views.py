from django.shortcuts import render, redirect , get_object_or_404
from .database_query import get_all_departments,get_all_locations,get_departments,designation_list_query,get_designations,location_list_query,employee_list_query,user_list_query
from datetime import datetime
from django.db import connection
from django.contrib import messages
from .models import Department,Designation,Location, Skills,Employee,User # Import your model
from django.contrib.auth.decorators import login_required
from .forms import DepartmentForm,DesignationForm,LocationForm,EmployeeForm,SkillFormSet,User_Form,User_Edit_Form,CustomLoginForm
import os
from django.conf import settings
from django.contrib.staticfiles import finders
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseServerError
from django.core.exceptions import ValidationError
import pandas as pd # type: ignore
import pandas as pd
from django.utils import timezone
import openpyxl
from openpyxl.styles import PatternFill
from django.http import HttpResponse
from io import BytesIO
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.core.files.storage import default_storage
from django.contrib.auth.hashers import make_password
from django.contrib.auth import authenticate, login,logout
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Image
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.units import inch
from django.core.mail import EmailMessage





def dashboard(request):
    return render(request, 'dashboard.html')


def indexpage(request):
    return render(request, 'index.html')

#----------------------------------------------USER ------------------------------------------------------------

def user_login(request):

    template_name = 'login.html'

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user_exist = User.objects.filter(username=username).exists()
       
        if user_exist:
           
            user = authenticate(request, username=username, password=password)
           
            if user is not None:
                if user.role == 'ADMIN' :
                    login(request, user)
           
                    return redirect('indexpage')
                
                elif user.role == 'VIEWER':
                    login(request, user)
                    return redirect('indexpage')
                
                else:
                    context = {'msg': 'Invalid Username or Password!'}
                    return render(request, template_name, context)
            else:
                
                context = {'msg': 'Password is incorrect!'}
                return render(request, template_name, context)

        else:
            context = {'msg': 'User Does Not exist'}
            return render(request, template_name, context)  
            
    return render(request, template_name)


def admin_logout(request):
    
    logout(request)
  
    return redirect(user_login)

#-------------------------DEPARTMENT------------------------------------------------------------------------

# ---------------------- LIST DEPARTMENT FUNCTION --------------------------------------


@login_required
def department_list(request):
    user_obj = User.objects.filter(id =request.user.id).first()
    role = user_obj.role
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Handle AJAX request for server-side processing
        return get_all_departments(request)
    
    # Handle normal page request
    
    return render(request, 'department_list.html',{
        'user': request.user ,
        'role': role # Make sure user is passed to the template context
    })

#--------------------- Add Department ----------------------------------------------------
@login_required
def department_add(request):
    
    if request.method == 'POST':
        department_name = request.POST.get('department_name')
        description = request.POST.get('description')
        created_at = datetime.now()
        created_by = request.user
        
        # created_by = User.objects.get(id=request.user.id)
        if department_name and description:
            Department.objects.create(
                department_name=department_name,
                description=description,
                created_at=created_at.now(),  
                created_by=created_by       
            )
            messages.success(request, 'Department added successfully!')
            return redirect('department_list') 
        else:
            messages.error(request, 'Both department name and description are required.')

    return render(request, 'department-add.html')
    
#---------------------------Edit Department-------------------------------------------------------------------------------#
@login_required    
def department_edit(request,department_id):

    department_instance = get_object_or_404(Department, department_id=department_id)
    updated_by = request.user
    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=department_instance)
        if form.is_valid():
            department = form.save(commit=False)
            department.updated_at = datetime.now()  
            department.updated_by = updated_by 
            department.save()
            messages.success(request, 'Department updated successfully!')
            return redirect('department_list')  #
        else:
            messages.error(request, 'Both department name and description are required.')

    else:
        form = DepartmentForm(instance=department_instance)
    return render(request, 'department-edit.html', {'form': form,'department': department_instance})

#---------------------------View  Department-------------------------------------------------------------------------------#

def department_view(request,department_id):
    department_instance = get_object_or_404(Department, department_id=department_id)
    return render(request, 'department-view.html',{'department': department_instance})

#---------------------------Delete  Department-------------------------------------------------------------------------------#

@login_required
def department_delete(request, department_id):
    
    department = get_object_or_404(Department, department_id=department_id)
    if request.method == 'POST':
        department.delete()
        messages.success(request, 'Department deleted successfully!')
        return redirect('department_list')  
    else:
        messages.error(request, 'Invalid request method.')
        return redirect('department_list')
    
#---------------------------Download  Department Template-------------------------------------------------------------------------------#
@login_required
def download_template(request):
    
    template_path = finders.find('template/department_template.xlsx')
    if not template_path:
        # Handle the case where the file is not found
        return HttpResponse("File not found", status=404)
    with open(template_path, 'rb') as template_file:
        response = HttpResponse(template_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="department_template.xlsx"'
        return response
    

#---------------------------Upload Department Details-------------------------------------------------------------------------------#
@login_required
def bulk_upload(request):
        
        # Check file type
        uploaded_file = request.FILES.get('upload_file')
        if not uploaded_file.name.endswith(('.xls', '.xlsx')):
            raise ValidationError("Invalid file format. Only .xls and .xlsx files are supported.")
        

        df = pd.read_excel(uploaded_file)
        if uploaded_file:
            for index, row in df.iterrows():
                    department_name = row.get('Department Name')
                    description = row.get('Description')
                    created_by = request.user
                    created_at =  timezone.now()
                    Department.objects.update_or_create(
                        department_name=department_name,  
                        defaults={
                            'description': description,
                            'created_at': created_at,
                            'created_by': created_by
                        }
                    ) 
            

            messages.success(request, 'Successfully Added !')
            return redirect('department_list')
        return render(request, 'department-add.html')
            
#---------------------------Export Department Details-------------------------------------------------------------------------------#

def export_departments(request):
    # Use the get_all_departments function to retrieve the department data
    departments = get_departments()
    for index, department in enumerate(departments, start=1):
        department['Sl.No'] = index
    # Convert the data to a DataFrame
    df = pd.DataFrame(departments)
    df = df[['Sl.No', 'department_name', 'description']]
    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=departments.xlsx'

    # Write the DataFrame to the response
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Departments')

    return response

#---------------------------END OF DEPARTMENT-------------------------------------------------------------------------------#

#######################################################################################################################################################

#---------------------------DESIGNATION-------------------------------------------------------------------------------#

# ---------------------- List Designation --------------------------------------

def designations_of_department(request):
    department_id = request.GET.get('department_id')
    designations = Designation.objects.filter(department=department_id).all()
    return JsonResponse(list(designations.values('designation_id', 'designation_name')), safe=False)
@csrf_exempt

def designation_list(request):
    user_obj = User.objects.filter(id =request.user.id).first()
    role = user_obj.role
    if request.method == "GET":
        template_name = 'designation_list.html'
        context ={
            'role':role
        }
        return render(request, template_name, context)

    if request.method == "POST":
        start_index = request.POST.get('start')
        page_length = request.POST.get('length')
        search_value = request.POST.get('search[value]')
        draw = request.POST.get('draw')
       
        des = designation_list_query(start_index, page_length, search_value, draw)
        return JsonResponse(des)

#--------------------- Add Designation ----------------------------------------------------
@login_required
def designation_add(request):
    
    if request.method == 'POST':
        designation_name = request.POST.get('designation_name')
        description = request.POST.get('description')
        created_at = datetime.now()
        created_by = request.user
        department_id = request.POST.get('department_id')
        
        if designation_name and description:
            Designation.objects.create(
                designation_name=designation_name,
                description=description,
                created_at=created_at,  
                created_by=created_by,
                department_id = department_id,   
            )
            messages.success(request, 'Designation added successfully!')
            return redirect('designation_list') 
        else:
            messages.error(request, 'Both Designation name and description are required.')
    
    departments = Department.objects.all().values('department_id', 'department_name')
    return render(request, 'designation_add.html',{'depts':departments})
    
#---------------------------Edit Designation-------------------------------------------------------------------------------#
@login_required   
def designation_edit(request, designation_id):
    
    designation_instance = get_object_or_404(Designation, pk=designation_id)
    departments = Department.objects.all()  # Get all departments for the dropdown
    current_department_id = designation_instance.department.pk if designation_instance.department else None
    current_department_name = designation_instance.department.department_name if designation_instance.department else None
    
    if request.method == 'POST':
        form = DesignationForm(request.POST, instance=designation_instance)
        if form.is_valid():
            # Save the form data
            designation = form.save(commit=False)
            designation.updated_at = timezone.now() 
            designation.updated_by = request.user  
            designation.save()
            
            messages.success(request, 'Designation updated successfully!')
            return redirect('designation_list') 
        else:
            messages.error(request, 'Both designation name and description are required.')

    else:
        form = DesignationForm(instance=designation_instance)

    
    return render(request, 'designation_edit.html', {'form': form,'designation': designation_instance,'departments': departments,'current_department_name': current_department_name,'current_department_id':current_department_id })



#---------------------------View  Designation-------------------------------------------------------------------------------#

def designation_view(request,designation_id):
    designation_instance = get_object_or_404(Designation, designation_id=designation_id)
    dept_obj = Department.objects.filter(department_id = designation_instance.department.department_id).first()
    dept_name = dept_obj.department_name
    return render(request, 'designation_view.html',{'designation': designation_instance,'dept_name':dept_name})

#---------------------------Delete  Designation-------------------------------------------------------------------------------#
@login_required
def designation_delete(request, designation_id):
    
    designation = get_object_or_404(Designation, designation_id=designation_id)
    if request.method == 'POST':
        designation.delete()
        messages.success(request, 'Designation deleted successfully!')
        return redirect('designation_list')  
    else:
        messages.error(request, 'Invalid request method.')
        return redirect('designation_list')
    

#---------------------------Download  Designation Template-------------------------------------------------------------------------------#

@login_required
def designation_download_template(request):
    
    departments = Department.objects.values_list('department_name', flat=True)
    if not departments:
        departments = ["No departments available"]  # Fallback if no departments exist

    # Create a new workbook and add worksheets
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"

    # Add column headers for data entry
    ws['A1'] = "Department Name"
    ws['B1'] = "Designation Name"
    ws['C1'] = "Description"

    

    # Add a new worksheet for the department list
    ws_list = wb.create_sheet(title="Department List")

    

    # Add department names to the new sheet
    for idx, dept in enumerate(departments, start=2):
        ws_list[f'A{idx}'] = dept

    # Save the workbook to a BytesIO object
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Serve the file as a downloadable response
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="designation_template.xlsx"'
    return response


#---------------------------Bulk Upload Designation -------------------------------------------------------------------------------#

@login_required
def designation_bulk_upload(request):
     
     if request.method == 'POST':
        uploaded_file = request.FILES.get('upload_file')
        
        if not uploaded_file or not uploaded_file.name.endswith(('.xls', '.xlsx')):
            messages.error(request, "Invalid file format. Only .xls and .xlsx files are supported.")
            return render(request, 'designation_add.html')

        try:
            df = pd.read_excel(uploaded_file)
        except Exception as e:
            messages.error(request, f"Error reading the file: {str(e)}")
            return render(request, 'designation_add.html')

        if df.empty:
            messages.error(request, 'The uploaded file is empty or does not contain valid data.')
            return render(request, 'designation_add.html')

        # Fetch existing department names for validation
        existing_departments = set(Department.objects.values_list('department_name', flat=True))

        # Track rows with warnings
        warnings = []

        for index, row in df.iterrows():
            department_name = row.get('Department Name')
            designation_name = row.get('Designation Name')
            description = row.get('Description')
            created_by = request.user
            created_at = timezone.now()

            # Skip rows with missing or invalid data
            if pd.isna(department_name) or not department_name.strip():
                warnings.append(f"Row {index + 1}: Department name is missing or invalid. This row will be ignored.")
                continue

            if pd.isna(designation_name) or not designation_name.strip():
                warnings.append(f"Row {index + 1}: Designation name is missing or invalid. This row will be ignored.")
                continue

            if pd.isna(description) or not description.strip():
                warnings.append(f"Row {index + 1}: Description is missing or invalid. This row will be ignored.")
                continue

            # Check if the department is valid
            if department_name not in existing_departments:
                warnings.append(f"Row {index + 1}: Department '{department_name}' is not in the department list. This row will be ignored.")
                continue

            # Get department instance
            try:
                department_instance = Department.objects.get(department_name=department_name)
            except Department.DoesNotExist:
                warnings.append(f"Row {index + 1}: Department '{department_name}' not found. This row will be ignored.")
                continue

            # Update or create the designation
            try:
                print(f"Processing Row {index + 1}: Designation Name = {designation_name}, Department = {department_instance.department_name}")
                Designation.objects.update_or_create(
                    designation_name=designation_name,
                    department_id = department_instance.department_id,
                    defaults={
                        'description': description,
                        'created_at': created_at,
                        'created_by': created_by,
                          # Use 'id' to refer to the primary key
                    }
                )
            except Exception as e:
                warnings.append(f"An error occurred while processing row {index + 1}: {str(e)}")

        # Provide feedback to the user
        if warnings:
            for warning in warnings:
                messages.warning(request, warning)
        else:
            messages.success(request, 'Successfully added/updated designations!')

        return redirect('designation_list')

     return render(request, 'designation_add.html')

#---------------------------Export Designation Details-------------------------------------------------------------------------------#

def export_designations(request):
    # Use the get_all_designation function to retrieve the designation data
    designations = get_designations()
 
    for index, designation in enumerate(designations, start=1):
        designation['Sl.No'] = index
    # Convert the data to a DataFrame
    df = pd.DataFrame(designations)
    df = df[['Sl.No', 'designation_name', 'description','department_name']]
    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=designations.xlsx'

    # Write the DataFrame to the response
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Designations')

    return response

##########################################################################################################################################

#--------------------------------------------LOCATIONS---------------------------------------------------------------------------------------

# ---------------------- List Locations --------------------------------------

@csrf_exempt
def location_list(request):
    user_obj = User.objects.filter(id =request.user.id).first()
    role = user_obj.role
    if request.method == "GET":
        template_name = 'location_list.html'
        context = {
            'role': role
        }
        return render(request, template_name, context)

    if request.method == "POST":
        start_index = request.POST.get('start')
        page_length = request.POST.get('length')
        search_value = request.POST.get('search[value]')
        draw = request.POST.get('draw')
       
        loc = location_list_query(start_index, page_length, search_value, draw)
       
        return JsonResponse(loc)

#--------------------- Add Location ----------------------------------------------------
@login_required
def location_add(request):
    
    if request.method == 'POST':
        location_name = request.POST.get('location_name')
        description = request.POST.get('description')
        created_at = datetime.now()
        created_by = request.user
        if location_name and description:
            Location.objects.create(
                location_name=location_name,
                description=description,
                created_at=created_at,  
                created_by=created_by       
            )
            messages.success(request, 'Location added successfully!')
            return redirect('location_list') 
        else:
            messages.error(request, 'Both Location name and description are required.')

    return render(request, 'location_add.html')



#---------------------------Edit Location-------------------------------------------------------------------------------#
@login_required  
def location_edit(request,location_id):
    
    location_instance = get_object_or_404(Location, location_id=location_id)
    if request.method == 'POST':
        form = LocationForm(request.POST, instance=location_instance)
        if form.is_valid():
            location = form.save(commit=False)
            location.updated_at = datetime.now()  
            location.updated_by = request.user  
            location.save()
            messages.success(request, 'Location updated successfully!')
            return redirect('location_list')  
        else:
            messages.error(request, 'Both Location name and description are required.')

    else:
        form = LocationForm(instance=location_instance)
    return render(request, 'location_edit.html', {'form': form,'location': location_instance})


#---------------------------View  Location-------------------------------------------------------------------------------#

def location_view(request,location_id):
    location_instance = get_object_or_404(Location, location_id=location_id)
    return render(request, 'location_view.html',{'location': location_instance})


#---------------------------Delete  Location-------------------------------------------------------------------------------#

@login_required
def location_delete(request, location_id):
    
    location = get_object_or_404(Location, location_id=location_id)
    if request.method == 'POST':
        location.delete()
        messages.success(request, 'Location deleted successfully!')
        return redirect('location_list')  
    else:
        messages.error(request, 'Invalid request method.')
        return redirect('location_list')
    download_location_template

#---------------------------Download  Location Template-------------------------------------------------------------------------------#
@login_required
def download_location_template(request):
    
    template_path = finders.find('template/location_template.xlsx')
    if not template_path:
        # Handle the case where the file is not found
        return HttpResponse("File not found", status=404)
    with open(template_path, 'rb') as template_file:
        response = HttpResponse(template_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="location_template.xlsx"'
        return response
    

#---------------------------Upload location Details-------------------------------------------------------------------------------#
@login_required
def location_bulk_upload(request):
        
        # Check file type
        uploaded_file = request.FILES.get('upload_file')
        if not uploaded_file.name.endswith(('.xls', '.xlsx')):
            raise ValidationError("Invalid file format. Only .xls and .xlsx files are supported.")
            

        df = pd.read_excel(uploaded_file)
        if uploaded_file:
            for index, row in df.iterrows():
                    location_name = row.get('Location Name')
                    description = row.get('Description')
                    created_by = request.user
                    created_at =  timezone.now()
                    Location.objects.update_or_create(
                        location_name=location_name,  
                        defaults={
                            'description': description,
                            'created_at': created_at,
                            'created_by': created_by
                        }
                    ) 
            

            messages.success(request, 'Successfully Added !')
            return redirect('location_list')
        return render(request, 'location_list.html')
            
#---------------------------Export Location Details-------------------------------------------------------------------------------#

def export_locations(request):
    # Use the get_all_locations function to retrieve the department data
    locations = get_all_locations()
    for index, location in enumerate(locations, start=1):
        location['Sl.No'] = index
    # Convert the data to a DataFrame
    df = pd.DataFrame(locations)
    df = df[['Sl.No', 'location_name', 'description']]
    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Locations.xlsx'

    # Write the DataFrame to the response
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Locations')

    return response

#-------------------------EMPLOYEE------------------------------------------------------------------------

# ---------------------- LIST EMPLOYEE FUNCTION --------------------------------------

@csrf_exempt
def employee_list(request):
    user_obj = User.objects.filter(id =request.user.id).first()
    role = user_obj.role
    if request.method == "GET":
        template_name = 'employee_list.html'
        context = {
            'role':role
        }
        return render(request, template_name,context)

    if request.method == "POST":
        start_index = request.POST.get('start')
        page_length = request.POST.get('length')
        search_value = request.POST.get('search[value]')
        draw = request.POST.get('draw')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
       
        emp = employee_list_query(start_index, page_length, search_value, draw, start_date, end_date)
       
        return JsonResponse(emp)
    

# ---------------------- Add Employee --------------------------------------
@login_required
def employee_add(request):
    
    form = EmployeeForm
    formset = SkillFormSet(queryset=Skills.objects.none())
    template_name = 'employee_add.html'
    context = {'form': form, 'formset': formset}
   
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES)

        formset = SkillFormSet(request.POST, queryset=Skills.objects.none())
        if form.is_valid() and formset.is_valid():
            data = form.save(commit=False)
            data.created_by = request.user
            data.save()
            
            
            for skill_form in formset:
                skill = skill_form.save(commit=False)
                skill.employee = data
                skill.save()
            messages.success(request, 'Employee Added Successfully', 'alert-success')
            return redirect('employee_list')
            
        else:
            messages.error(request, 'Data is not valid.', extra_tags='alert-danger')
            context = {'form': form,'formset': formset}
            return render(request, template_name, context)
    else :
        return render(request, template_name, context)
    
# ---------------------- Edit Employee --------------------------------------

@login_required
def employee_edit(request,employee_id):
    
    template_name = 'employee_edit.html'
    employee_obj = Employee.objects.get(employee_id = employee_id)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES, instance=employee_obj)
        formset = SkillFormSet(request.POST, queryset=Skills.objects.filter(employee=employee_obj))
        if form.is_valid() :
         
            employee = form.save(commit=False)
            employee.updated_at = datetime.now()
            employee.updated_by = request.user
            employee.save()
           
            if formset.is_valid():
                for i in formset:
                  
                    skill = i.save()
                    skill.employee = employee
                    skill.save()
             
            messages.success(request, 'Employee Successfully Updated.', 'alert-success')
            return redirect('employee_list')
        else:
          
            messages.error(request, 'Data is not valid.', 'alert-danger')
    else:
        form = EmployeeForm(instance=employee_obj)
        formset = SkillFormSet(queryset=Skills.objects.filter(employee=employee_obj))
    
    context = {'form': form, 'formset': formset, 'employee_obj': employee_obj}
    return render(request, template_name, context)

# ---------------------- View Employee --------------------------------------
@login_required
def employee_view(request,employee_id):

    employee = get_object_or_404(Employee,employee_id=employee_id)
    department=employee.department.department_name
    designation=employee.designation.designation_name
    location=employee.location.location_name
    skills = Skills.objects.filter(employee=employee)
    context = {
        
        'employee': employee,
        'department':department,
        'location':location,
        'designation':designation,
        'skills': skills,
    }
    
    return render(request, 'employee_view.html', context)


#---------------------------Delete  Employee-------------------------------------------------------------------------------#
@login_required
def employee_delete(request, employee_id):
    
    employee = get_object_or_404(Employee, employee_id=employee_id)
    if request.method == 'POST':
        employee.delete()
        messages.success(request, 'Employee deleted successfully!')
        return redirect('employee_list')  
    else:
        messages.error(request, 'Invalid request method.')
        return redirect('employee_list')
    
#---------------------------Export Employee Details-------------------------------------------------------------------------------#
@login_required
def export_employee(request):
    employee = Employee.objects.all()
    data = []

    for index, employee in enumerate(employee, start=1):
        photo_url = employee.photo.url if employee.photo and default_storage.exists(employee.photo.name) else ''
        skills = [skill.skill_name or '' for skill in employee.skills.all()]
        data.append({
            'Sl.No': index,
            'Employee No': employee.employee_no,
            'Join Date': employee.join_date,
            'Name': employee.name,
            'Phone': employee.phone,
            'Address': employee.address,
            'Emp Start Date': employee.emp_start_date,
            'Emp End Date': employee.emp_end_date,
            'Photo': photo_url,
            'Status': employee.status,
            'Department': employee.department.department_name if employee.department else '',
            'Designation': employee.designation.designation_name if employee.designation else '',
            'Location': employee.location.location_name if employee.location else '',
            'Skills':  ', '.join(skills),
            
        })

    df = pd.DataFrame(data)

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employee.xlsx'

    # Write the DataFrame to the response
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='employee')

    return response


#---------------------------Download  Employee Template-------------------------------------------------------------------------------#

@login_required
def download_employee_template(request):
    
     # Define the headings for the Employee template
    employee_headers = [
        'Join Date',
        'Employee No',
        'Name',
        'Phone',
        'Address',
        'Employee Start Date',
        'Employee End Date',
        'Photo',
        'Status',
        'Department',
        'Designation',
        'Location'
    ]
    # Define the headings for the Skills template
    skills_headers = [
        'Employee Number',
        'Skill Name',
        'Description'
    ]
    # Create an Excel writer object
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Create a DataFrame with the headers for Employees
        df_employees = pd.DataFrame(columns=employee_headers)
        df_employees.to_excel(writer, sheet_name='Employees', index=False)

        # Create a DataFrame with the headers for Skills
        df_skills = pd.DataFrame(columns=skills_headers)
        df_skills.to_excel(writer, sheet_name='Skills', index=False)

    # Prepare the HTTP response
    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employee_template.xlsx'
    
    return response

#---------------------------Upload Employee Details-------------------------------------------------------------------------------#

@login_required
def bulk_upload_employees(request):
    
    if 'file' not in request.FILES:
        return HttpResponse("No file uploaded", status=400)
    
    file = request.FILES['file']
    
    try:
        # Read the Excel file
        excel_data = pd.ExcelFile(file)
        
        # Read employee data
        df_employees = pd.read_excel(excel_data, sheet_name='Employees')
        # Read skills data
        df_skills = pd.read_excel(excel_data, sheet_name='Skills')
        
        # Fetch existing departments, designations, and locations
        departments = {dep.department_name: dep for dep in Department.objects.all()}
        designations = {desg.designation_name: desg for desg in Designation.objects.all()}
        locations = {loc.location_name: loc for loc in Location.objects.all()}
        # Iterate over employee data
        for index, row in df_employees.iterrows():
            print(f"Processing row {index}: {row}")  # Debugging
            # Handle missing or incorrect data types
            department = departments.get(row['Department'])
            designation = designations.get(row['Designation'])
            location = locations.get(row['Location'])
            created_at = timezone.now()
            created_by = request.user
            
            if department and designation and location:
                created_at = timezone.now()
                created_by = request.user
                try:
                    # Convert photo to None if it's NaN
                    photo = row['Photo'] if pd.notna(row['Photo']) else None
                    
                    # Create or update employee
                    employee, created = Employee.objects.update_or_create(      
                        employee_no=row['Employee No'],
                        defaults={
                            'join_date': pd.to_datetime(row['Join Date'], errors='coerce'),
                            'name': row['Name'],
                            'phone': row['Phone'],
                            'address': row['Address'],
                            'emp_start_date': pd.to_datetime(row['Employee Start Date'], errors='coerce'),
                            'emp_end_date': pd.to_datetime(row['Employee End Date'], errors='coerce'),
                            'photo': photo,  # Ensure proper handling of file paths
                            'status': row['Status'],
                            'department': department,
                            'designation': designation,
                            'location': location,
                            'created_at' : created_at,  
                            'created_by' : created_by 
                        }
                    )
                    print(f"Employee {'created' if created else 'updated'}: {employee}")
                    # Now handle skills associated with this employee
                    skills_for_employee = df_skills[df_skills['Employee Number'] == row['Employee No']]
                    
                    for _, skill_row in skills_for_employee.iterrows():
                        Skills.objects.update_or_create(
                            employee=employee,
                            skill_name=skill_row['Skill Name'],
                            defaults={
                                'description': skill_row['Description']
                            }
                        )
                
                except Exception as e:
                    # Log the error and continue processing other rows
                    print(f"Error processing employee row {index}: {e}")
        
        return render(request, 'employee_list.html')
    
    except Exception as e:
        return HttpResponse(f"Error processing file: {e}", status=500)
    
#---------------------------Filter Employees with start and end date-------------------------------------------------------------------------------#

@csrf_exempt
def filtered_employees(request):
    if request.method == "GET":
        template_name = 'employee_list.html'
        return render(request, template_name)

    if request.method == "POST":
        start_index = request.POST.get('start')
        page_length = request.POST.get('length')
        search_value = request.POST.get('search[value]')
        draw = request.POST.get('draw')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        emp = employee_list_query(start_index, page_length, search_value, draw, start_date, end_date)

        return JsonResponse(emp)
   
#---------------------------Export selected Employees -------------------------------------------------------------------------------#

def export_selected_employees(request):
    if request.method == 'POST':
        # Retrieve employee IDs from POST data
        employee_ids = request.POST.get('employee_ids')
        employee_ids = json.loads(employee_ids)

        # Filter employees based on IDs
        employees = Employee.objects.filter(employee_id__in=employee_ids)
        data = []

        for index, employee in enumerate(employees, start=1):
            photo_url = employee.photo.url if employee.photo and default_storage.exists(employee.photo.name) else ''

            data.append({
                'Sl.No': index,
                'Employee No': employee.employee_no,
                'Join Date': employee.join_date,
                'Name': employee.name,
                'Phone': employee.phone,
                'Address': employee.address,
                'Emp Start Date': employee.emp_start_date,
                'Emp End Date': employee.emp_end_date,
                'Photo': photo_url,
                'Status': employee.status,
                'Department': employee.department.department_name if employee.department else '',
                'Designation': employee.designation.designation_name if employee.designation else '',
                'Location': employee.location.location_name if employee.location else '',
                'Skills': ', '.join([skill.skill_name for skill in employee.skills.all()]),
            })

        df = pd.DataFrame(data)
        # Create an HTTP response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=selected_employees.xlsx'

        # Write the DataFrame to the response
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='employees')

        return response
    else:
        return HttpResponse(status=405)  

#------------------------------------ Download PDF of an employee --------------------------------------------------------------------
def generate_pdf(employee):
    # Create a PDF buffer to hold the PDF data
    buffer = BytesIO()

    # Create the PDF document with the specified page size
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    # Add employee photo if available
    if employee.photo:  # Assuming you have a `photo` field in your Employee model
        try:
            photo_path = employee.photo.path
            img = Image(photo_path, width=2*inch, height=2*inch)
            elements.append(img)
            elements.append(Spacer(1, 12))
        except Exception as e:
            # Handle the case where the photo cannot be loaded
            print(f"Error loading image: {e}")

    # Retrieve related data
    department = employee.department.department_name
    designation = employee.designation.designation_name
    location = employee.location.location_name
    skills = Skills.objects.filter(employee=employee)

    # Get styles for text formatting
    styles = getSampleStyleSheet()

    # Add employee details to the PDF
    elements.append(Paragraph(f"Employee No: {employee.employee_no}", styles['Title']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Name: {employee.name}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Phone: {employee.phone}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Address: {employee.address}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Employee Join Date: {employee.join_date}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Employee Start Date: {employee.emp_start_date}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Employee End Date: {employee.emp_end_date}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Status: {employee.status}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Department: {department}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Designation: {designation}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    elements.append(Paragraph(f"Location: {location}", styles['Normal']))
    elements.append(Spacer(1, 12))  # Add space between elements

    # Add skills to the PDF, handling the case where there are no skills
    if skills.exists():
        skills_list = ', '.join(skill.skill_name for skill in skills)
    else:
        skills_list = 'No skills recorded'

    elements.append(Paragraph(f"Skills: {skills_list}", styles['Normal']))

    # Build the PDF document
    doc.build(elements)

    # Get the PDF data from the buffer
    pdf = buffer.getvalue()
    buffer.close()
    
    return pdf

def employee_pdf_download(request, employee_id):
    # Get the employee object or return a 404 error if not found
    employee = get_object_or_404(Employee, employee_id=employee_id)
    
    # Generate the PDF
    pdf = generate_pdf(employee)
    
    # Return the PDF as an HTTP response with appropriate headers
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="employee_{employee.employee_no}.pdf"'
    
    # Send the PDF as an email attachment

    return response

def mail_pdf(request,employee_id):
     # Ensure the request method is POST
    if request.method != 'POST':
        return HttpResponseBadRequest("Invalid request method. Use POST.")

    # Get the employee object or return a 404 error if not found
    employee = get_object_or_404(Employee, employee_id=employee_id)
    
    # Generate the PDF
    pdf = generate_pdf(employee)
    
    # Get email address from POST data
    recipient_email = request.POST.get('email')
    if not recipient_email:
        return HttpResponseBadRequest("No email address provided.")
    
    # Define the email subject and body
    subject = f"Employee Report for {employee.name}"
    body = f"Dear {employee.name},\n\nPlease find attached the PDF report containing your details.\n\nBest regards,\nYour Company"
    
    try:
        # Create an email message
        email = EmailMessage(
            subject,
            body,
            'your_email@example.com',  # Replace with your sender email address
            [recipient_email],  # Replace with the employee's email address
        )
        
        # Attach the PDF
        email.attach(f"employee_{employee.employee_no}.pdf", pdf, 'application/pdf')
        
        # Send the email
        email.send()

        # Return a success response
        messages.success(request, 'Email sent successfully.','alert-success')
        return redirect('employee_list')

    except Exception as e:
        # Log the exception and return an error response
        print(f"Error sending email: {e}")
        return HttpResponseServerError("An error occurred while sending the email.")
#-------------------------------------- ACCOUNTS ------------------------------------------------------------------

#------------------------------------ User List --------------------------------------------------------------------

def user_list(request):

    user_obj = User.objects.filter(id =request.user.id).first()
    role = user_obj.role
    if request.method == "GET":
        template_name = 'user_list.html'
        context = {
            'role': role
        }
        return render(request, template_name,context)

    if request.method == "POST":      
        start_index = request.POST.get('start')
        page_length = request.POST.get('length')
        search_value = request.POST.get('search[value]')
        draw = request.POST.get('draw')
       
        des = user_list_query(start_index, page_length, search_value, draw)
       
        return JsonResponse(des)

#------------------------------------ User Add --------------------------------------------------------------------

@login_required
def user_add(request):
    
    form = User_Form
   
    template_name = 'user_add.html'
    
    context = {'form': form}
    if request.method == 'POST':
        form = User_Form(request.POST, request.FILES)
        
        if form.is_valid() :
            data = form.save(commit=False)
           
            passw = data.password
            passw = make_password(passw)
            data.password = passw
           
            data.save()
            messages.success(request, 'User Added Successfully', 'alert-success')
            return redirect('user_list')
        else:
            messages.error(request, 'Data is not valid.', 'alert-danger')
            context = {'form': form,}
            return render(request, template_name, context)
    else:
        return render(request, template_name, context)
    
# -------------------------------- User Edit ---------------------------------------------
@login_required
def user_edit(request, id):
    
    template_name = 'user_edit.html'
   
    
    user_obj = User.objects.get(id=id)
    role = user_obj.role
    form = User_Edit_Form(instance=user_obj)
    
    context = {'form': form,}
    if request.method == 'POST':
        form = User_Edit_Form(request.POST, request.FILES, instance=user_obj)
       
        if form.is_valid():
            data = form.save(commit=False)
           
            data.save()
            messages.success(request, 'User Updated Successfully', 'alert-success')
            return redirect('user_list')
        else:
            
            context = {'form': form,}
            print(form.errors)
            messages.error(request, 'Data is not valid.', 'alert-danger')
            return render(request, template_name, context)
    else:
        return render(request, template_name, context)
    

#------------------------------------ View User ---------------------------------------------------
@login_required
def user_view(request,id):

    user = get_object_or_404(User,id=id)
    
    context = {
        'user': user
    }
    
    return render(request, 'user_view.html', context)

   
# --------------------------------- Delete User --------------------------------------------------------
@login_required
def user_delete(request, id):
    
    user = User.objects.get(id=id)
    
    user.delete()
    messages.success(request, 'User Deleted Successfully', 'alert-success')
    return redirect('user_list')
    
   
# -------------------------- Download Template ----------------------------------------------------------

@login_required
def download_user_template(request):
    
    user_headers = [
            'Username',
            'First Name',
            'Last Name',
            'Email',
            'Password',
            'Role'
        ]
        
        # Create an Excel writer object
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Create a DataFrame with the headers for Users
        df_users = pd.DataFrame(columns=user_headers)
        df_users.to_excel(writer, sheet_name='Users', index=False)

    # Prepare the HTTP response
    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=user_template.xlsx'
        
    return response


# ---------------------------- Bulk Upload User -----------------------------------------------------------------
@login_required
def user_bulk_upload(request):
    
    if 'file' not in request.FILES:
        return HttpResponse("No file uploaded", status=400)
    
    file = request.FILES['file']
    
    try:
        # Read the Excel file
        excel_data = pd.ExcelFile(file)
        
        # Read user data
        df_users = pd.read_excel(excel_data, sheet_name='Users')
        
        
        # Iterate over user data
        for index, row in df_users.iterrows():
            
            try:
                # Handle missing or incorrect data
                username = row['Username']
                first_name = row['First Name']
                last_name = row['Last Name']
                email = row['Email']
                password=row['Password'] 
                role = row['Role']
                
                if pd.isna(username) or pd.isna(email) or pd.isna(password):
                    # Skip rows with missing mandatory fields
                    print(f"Skipping row {index} due to missing mandatory fields.")
                    continue
                
                if role not in ['Admin', 'Viewer','ADMIN','VIEWER']:
                    # Skip rows with invalid roles
                    print(f"Skipping row {index} due to invalid role: {role}")
                    continue
                
                # Create or update user
                user, created = User.objects.update_or_create(
                    username=username,
                    defaults={
                        'first_name': first_name,
                        'last_name': last_name,
                        'email': email,
                        'role': role,
                        'password': make_password(password),
                        

                    }
                )
                
              
            except Exception as e:
                # Log the error and continue processing other rows
                print(f"Error processing user row {index}: {e}")
        
        return render(request, 'user_list.html')
    
    except Exception as e:
        return HttpResponse(f"Error processing file: {e}", status=500)
    
# -------------------------------- Export Users -----------------------------------------------

@login_required
def export_user_details(request):
    # Fetch user data from the database
    users = User.objects.all().values(
        'username',
        'first_name',
        'last_name',
        'email',
        'role'
    )
    
    # Convert data to a DataFrame
    df_users = pd.DataFrame(users)
    
    # Define the response as an Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_users.to_excel(writer, sheet_name='Users', index=False)
    
    # Prepare HTTP response
    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=user_details.xlsx'
    
    return response
# -------------------------------- Edit Profile -----------------------------------------------
@login_required
def edit_profile(request):
    template_name = 'profile_edit.html'
    user_id =request.user.id
    user_obj = User.objects.get(id=user_id)
    
    form = User_Edit_Form(instance=user_obj)
    context = {'form': form,}
    if request.method == 'POST':
        form = User_Edit_Form(request.POST, request.FILES, instance=user_obj)
       
        if form.is_valid():
            data = form.save(commit=False)
           
            data.save()
            messages.success(request, 'User Updated Successfully', 'alert-success')
            return redirect('indexpage')
        else:
            
            context = {'form': form,}
            print(form.errors)
            messages.error(request, 'Data is not valid.', 'alert-danger')
            return render(request, template_name, context)
    else:
        return render(request, template_name, context)